from flask import Blueprint, render_template, request, redirect, session, send_file
import sqlite3
import io
from datetime import datetime
import pytz
from collections import defaultdict, deque
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font

admin_bp = Blueprint("admin", __name__)
DB_PATH = 'checkin.db'
tz = pytz.timezone("Asia/Taipei")

ADMIN_USERNAME = 'admin'
ADMIN_PASSWORD = 'admin'

@admin_bp.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        if request.form["username"] == ADMIN_USERNAME and request.form["password"] == ADMIN_PASSWORD:
            session["admin"] = True
            return redirect("/admin/dashboard")
        else:
            return render_template("admin_login.html", error="帳號或密碼錯誤")
    return render_template("admin_login.html")

@admin_bp.route("/logout")
def logout():
    session.clear()
    return redirect("/admin/login")

@admin_bp.route("/dashboard")
def dashboard():
    if not session.get("admin"):
        return redirect("/admin/login")
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    cursor.execute("SELECT employee_id, name, bind_time FROM users")
    users = cursor.fetchall()
    cursor.execute("SELECT DISTINCT DATE(timestamp) FROM checkins ORDER BY timestamp DESC")
    checkin_dates = [r[0] for r in cursor.fetchall()]
    cursor.execute("SELECT DISTINCT DATE(timestamp) FROM location_logs ORDER BY timestamp DESC")
    location_dates = [r[0] for r in cursor.fetchall()]
    checkin_months = sorted({d[:7] for d in checkin_dates}, reverse=True)
    location_months = sorted({d[:7] for d in location_dates}, reverse=True)
    conn.close()
    return render_template("admin_dashboard.html", users=users, checkin_dates=checkin_dates,
                           location_dates=location_dates, checkin_months=checkin_months,
                           location_months=location_months)

@admin_bp.route("/delete_user/<employee_id>")
def delete_user(employee_id):
    if not session.get("admin"):
        return redirect("/admin/login")
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    cursor.execute("DELETE FROM users WHERE employee_id=?", (employee_id,))
    cursor.execute("DELETE FROM user_states WHERE temp_employee_id=?", (employee_id,))
    conn.commit()
    conn.close()
    return redirect("/admin/dashboard")

@admin_bp.route("/export_checkins_excel", methods=["POST"])
def export_checkins_excel():
    if not session.get("admin"):
        return redirect("/admin/login")
    daterange = request.form.get("daterange") or ""
    if " - " in daterange:
        start_date, end_date = [d.strip() for d in daterange.split(" - ")]
    elif len(daterange) == 7:
        start_date = f"{daterange}-01"
        end_date = f"{daterange}-31"
    else:
        return "無效的日期格式", 400

    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    cursor.execute('''SELECT employee_id, name, timestamp, check_type FROM checkins
                      WHERE DATE(timestamp) BETWEEN ? AND ?
                      ORDER BY employee_id, timestamp''', (start_date, end_date))
    records = cursor.fetchall()
    conn.close()

    grouped = defaultdict(deque)
    for emp_id, name, ts, ctype in records:
        dt = tz.localize(datetime.strptime(ts, "%Y-%m-%d %H:%M:%S"))
        grouped[(emp_id, name)].append((dt, ctype))

    structured = defaultdict(lambda: defaultdict(lambda: {"上班": "", "下班": ""}))
    all_dates = set()
    for key, events in grouped.items():
        i = 0
        while i < len(events):
            dt, ctype = events[i]
            if ctype == "上班":
                date_str = dt.strftime("%Y/%m/%d")
                structured[key][date_str]["上班"] = dt.strftime("%H:%M")
                if i + 1 < len(events):
                    next_dt, next_type = events[i + 1]
                    if next_type == "下班":
                        structured[key][date_str]["下班"] = next_dt.strftime("%H:%M")
                        i += 1
                all_dates.add(date_str)
            i += 1

    sorted_dates = sorted(all_dates)
    employees = list(structured.keys())
    groups = [employees[i:i + 3] for i in range(0, len(employees), 3)]

    wb = Workbook()
    wb.remove(wb.active)
    for idx, group in enumerate(groups):
        ws = wb.create_sheet(title=f"第{idx + 1}頁")
        for i, (emp_id, name) in enumerate(group):
            col = i * 4 + 1
            ws.merge_cells(start_row=1, start_column=col, end_row=1, end_column=col + 2)
            ws.cell(row=1, column=col, value=f"工號：{emp_id} 姓名：{name}").font = Font(bold=True)
            ws.cell(row=2, column=col, value="日期").alignment = Alignment(horizontal="center")
            ws.cell(row=2, column=col+1, value="上班").alignment = Alignment(horizontal="center")
            ws.cell(row=2, column=col+2, value="下班").alignment = Alignment(horizontal="center")
            for row_idx, date in enumerate(sorted_dates, start=3):
                ws.cell(row=row_idx, column=col, value=date)
                ws.cell(row=row_idx, column=col+1, value=structured[(emp_id, name)][date]["上班"])
                ws.cell(row=row_idx, column=col+2, value=structured[(emp_id, name)][date]["下班"])

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    filename = f"打卡紀錄_{start_date}_to_{end_date}.xlsx"
    return send_file(out, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name=filename)

@admin_bp.route("/export_locations_excel", methods=["POST"])
def export_locations_excel():
    if not session.get("admin"):
        return redirect("/admin/login")
    daterange = request.form.get("daterange") or ""
    if " - " in daterange:
        start_date, end_date = [d.strip() for d in daterange.split(" - ")]
    elif len(daterange) == 7:
        start_date = f"{daterange}-01"
        end_date = f"{daterange}-31"
    else:
        return "無效的日期格式", 400

    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    cursor.execute('''SELECT employee_id, name, timestamp, latitude, longitude FROM location_logs
                      WHERE DATE(timestamp) BETWEEN ? AND ?
                      ORDER BY employee_id, timestamp''', (start_date, end_date))
    records = cursor.fetchall()
    conn.close()

    data = defaultdict(lambda: defaultdict(lambda: {"lat": "", "lng": ""}))
    all_dates = set()
    for emp_id, name, ts, lat, lng in records:
        dt = tz.localize(datetime.strptime(ts, "%Y-%m-%d %H:%M:%S"))
        date_str = dt.strftime("%Y/%m/%d")
        data[(emp_id, name)][date_str] = {"lat": lat, "lng": lng}
        all_dates.add(date_str)

    sorted_dates = sorted(all_dates)
    employees = list(data.keys())
    groups = [employees[i:i + 3] for i in range(0, len(employees), 3)]

    wb = Workbook()
    wb.remove(wb.active)
    for idx, group in enumerate(groups):
        ws = wb.create_sheet(title=f"定位{idx + 1}")
        for i, (emp_id, name) in enumerate(group):
            col = i * 4 + 1
            ws.merge_cells(start_row=1, start_column=col, end_row=1, end_column=col + 2)
            ws.cell(row=1, column=col, value=f"工號：{emp_id} 姓名：{name}").font = Font(bold=True)
            ws.cell(row=2, column=col, value="日期").alignment = Alignment(horizontal="center")
            ws.cell(row=2, column=col+1, value="緯度").alignment = Alignment(horizontal="center")
            ws.cell(row=2, column=col+2, value="經度").alignment = Alignment(horizontal="center")
            for row_idx, date in enumerate(sorted_dates, start=3):
                ws.cell(row=row_idx, column=col, value=date)
                ws.cell(row=row_idx, column=col+1, value=data[(emp_id, name)][date]["lat"])
                ws.cell(row=row_idx, column=col+2, value=data[(emp_id, name)][date]["lng"])

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    filename = f"定位紀錄_{start_date}_to_{end_date}.xlsx"
    return send_file(out, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name=filename)

@admin_bp.route("/clear_data")
def clear_data():
    if not session.get("admin"):
        return redirect("/admin/login")
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    cursor.execute("DELETE FROM checkins;")
    cursor.execute("DELETE FROM location_logs;")
    conn.commit()
    conn.close()
    return "\u2705 所有打卡與定位紀錄已清空"
